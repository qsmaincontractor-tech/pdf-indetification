"""
Test cases for Excel export (utils/excel_export.py).

Tests:
- Export creates a valid Excel file
- Sheet names are correct
- PDF File List sheet has correct data  
- PDF Page List sheet has correct data with extracted columns
- Column headers are present
- Empty project exports cleanly
"""

import os
import pytest
from openpyxl import load_workbook

from models.data_models import ProjectData, PDFFileInfo, PageData, ExtractedDataColumn
from utils.excel_export import export_to_excel


@pytest.fixture
def export_project():
    """Create a project with data suitable for export testing."""
    project = ProjectData()
    project.add_column("Title")
    project.add_column("Date")
    
    # File 1 with 2 pages
    file1 = PDFFileInfo(
        file_name="report.pdf",
        file_path="C:/docs/report.pdf",
        num_pages=2,
        file_size=50000,
    )
    page1 = PageData(page_number=0)
    page1.extracted_data = {"Title": "Annual Report", "Date": "2025-01-01"}
    page2 = PageData(page_number=1)
    page2.extracted_data = {"Title": "Summary", "Date": "2025-01-02"}
    file1.pages = [page1, page2]
    
    # File 2 with 1 page
    file2 = PDFFileInfo(
        file_name="memo.pdf",
        file_path="C:/docs/memo.pdf",
        num_pages=1,
        file_size=10000,
    )
    page3 = PageData(page_number=0)
    page3.extracted_data = {"Title": "Memo", "Date": ""}
    file2.pages = [page3]
    
    project.pdf_files = [file1, file2]
    return project


class TestExcelExport:
    """Tests for the Excel export functionality."""
    
    def test_creates_file(self, temp_dir, export_project):
        """Test that the export creates an Excel file."""
        path = os.path.join(temp_dir, "output.xlsx")
        export_to_excel(export_project, path)
        assert os.path.exists(path)
    
    def test_sheet_names(self, temp_dir, export_project):
        """Test that the workbook has the correct sheet names."""
        path = os.path.join(temp_dir, "output.xlsx")
        export_to_excel(export_project, path)
        
        wb = load_workbook(path)
        assert "PDF File List" in wb.sheetnames
        assert "PDF Page List" in wb.sheetnames
        wb.close()
    
    def test_file_list_headers(self, temp_dir, export_project):
        """Test that PDF File List sheet has correct headers."""
        path = os.path.join(temp_dir, "output.xlsx")
        export_to_excel(export_project, path)
        
        wb = load_workbook(path)
        ws = wb["PDF File List"]
        headers = [ws.cell(row=1, column=i).value for i in range(1, 5)]
        assert headers == ["File Name", "File Path", "Number of Pages", "File Size (bytes)"]
        wb.close()
    
    def test_file_list_data(self, temp_dir, export_project):
        """Test that PDF File List sheet has correct data."""
        path = os.path.join(temp_dir, "output.xlsx")
        export_to_excel(export_project, path)
        
        wb = load_workbook(path)
        ws = wb["PDF File List"]
        
        # Row 2 = first file
        assert ws.cell(row=2, column=1).value == "report.pdf"
        assert ws.cell(row=2, column=3).value == 2
        assert ws.cell(row=2, column=4).value == 50000
        
        # Row 3 = second file
        assert ws.cell(row=3, column=1).value == "memo.pdf"
        wb.close()
    
    def test_page_list_headers(self, temp_dir, export_project):
        """Test that PDF Page List sheet has correct headers including custom columns."""
        path = os.path.join(temp_dir, "output.xlsx")
        export_to_excel(export_project, path)
        
        wb = load_workbook(path)
        ws = wb["PDF Page List"]
        headers = [ws.cell(row=1, column=i).value for i in range(1, 6)]
        assert headers == ["File Name", "File Path", "Page Number", "Title", "Date"]
        wb.close()
    
    def test_page_list_data(self, temp_dir, export_project):
        """Test that PDF Page List sheet has correct extracted data."""
        path = os.path.join(temp_dir, "output.xlsx")
        export_to_excel(export_project, path)
        
        wb = load_workbook(path)
        ws = wb["PDF Page List"]
        
        # Row 2 = file1, page1
        assert ws.cell(row=2, column=1).value == "report.pdf"
        assert ws.cell(row=2, column=3).value == 1  # 1-based page number
        assert ws.cell(row=2, column=4).value == "Annual Report"
        assert ws.cell(row=2, column=5).value == "2025-01-01"
        
        # Row 3 = file1, page2
        assert ws.cell(row=3, column=3).value == 2
        assert ws.cell(row=3, column=4).value == "Summary"
        
        # Row 4 = file2, page1
        assert ws.cell(row=4, column=1).value == "memo.pdf"
        assert ws.cell(row=4, column=4).value == "Memo"
        wb.close()
    
    def test_page_count(self, temp_dir, export_project):
        """Test that the page list has the correct number of rows."""
        path = os.path.join(temp_dir, "output.xlsx")
        export_to_excel(export_project, path)
        
        wb = load_workbook(path)
        ws = wb["PDF Page List"]
        # Header + 3 data rows
        assert ws.max_row == 4
        wb.close()
    
    def test_empty_project(self, temp_dir):
        """Test exporting an empty project."""
        project = ProjectData()
        path = os.path.join(temp_dir, "empty.xlsx")
        export_to_excel(project, path)
        
        assert os.path.exists(path)
        wb = load_workbook(path)
        assert "PDF File List" in wb.sheetnames
        assert "PDF Page List" in wb.sheetnames
        wb.close()
    
    def test_project_with_no_columns(self, temp_dir):
        """Test exporting a project with files but no custom columns."""
        project = ProjectData()
        file1 = PDFFileInfo(file_name="a.pdf", file_path="/a.pdf", num_pages=1, file_size=100)
        file1.pages = [PageData(page_number=0)]
        project.pdf_files = [file1]
        
        path = os.path.join(temp_dir, "no_cols.xlsx")
        export_to_excel(project, path)
        
        wb = load_workbook(path)
        ws = wb["PDF Page List"]
        headers = [ws.cell(row=1, column=i).value for i in range(1, 4)]
        assert headers == ["File Name", "File Path", "Page Number"]
        wb.close()
