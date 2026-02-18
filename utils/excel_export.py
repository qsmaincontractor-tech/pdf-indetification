"""
Excel export utilities for the PDF Text Extraction application.

Exports project data to Excel format with two sheets:
- PDF File List: file name, file path, number of pages, file size
- PDF Page List: file name, file path, page number, extracted data columns
"""

import os
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from models.data_models import ProjectData


def export_to_excel(project_data: ProjectData, output_path: str) -> None:
    """
    Export project data to an Excel file.
    
    Creates an Excel workbook with two sheets:
    1. PDF File List: Overview of all imported PDF files.
    2. PDF Page List: Page-level data with all extracted data columns.
    
    Args:
        project_data: The ProjectData to export.
        output_path: Path where the Excel file will be saved.
        
    Raises:
        IOError: If the file cannot be written.
        
    Example:
        >>> from models.data_models import ProjectData
        >>> project = ProjectData()
        >>> export_to_excel(project, "output.xlsx")
    """
    wb = Workbook()
    
    # -- Style definitions --
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    # ===== Sheet 1: PDF File List =====
    ws_files = wb.active
    ws_files.title = "PDF File List"
    
    file_headers = ["File Name", "File Path", "Number of Pages", "File Size (bytes)"]
    for col_idx, header in enumerate(file_headers, 1):
        cell = ws_files.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, pdf_file in enumerate(project_data.pdf_files, 2):
        values = [
            pdf_file.file_name,
            pdf_file.file_path,
            pdf_file.num_pages,
            pdf_file.file_size,
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws_files.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # Auto-fit column widths for file list
    for col_idx in range(1, len(file_headers) + 1):
        max_length = len(file_headers[col_idx - 1])
        for row in ws_files.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws_files.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 60)
    
    # ===== Sheet 2: PDF Page List =====
    ws_pages = wb.create_sheet(title="PDF Page List")
    
    column_names = project_data.get_column_names()
    page_headers = ["File Name", "File Path", "Page Number"] + column_names
    
    for col_idx, header in enumerate(page_headers, 1):
        cell = ws_pages.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    row_idx = 2
    for pdf_file in project_data.pdf_files:
        for page in pdf_file.pages:
            values = [
                pdf_file.file_name,
                pdf_file.file_path,
                page.page_number + 1,  # 1-based page number for display
            ]
            for col_name in column_names:
                values.append(page.extracted_data.get(col_name, ""))
            
            for col_idx, value in enumerate(values, 1):
                cell = ws_pages.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
            
            row_idx += 1
    
    # Auto-fit column widths for page list
    for col_idx in range(1, len(page_headers) + 1):
        max_length = len(page_headers[col_idx - 1])
        for row in ws_pages.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws_pages.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 60)
    
    # Save
    wb.save(output_path)
